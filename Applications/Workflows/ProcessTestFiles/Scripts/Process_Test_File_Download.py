'''

@ Author - Ayushi Pawar
@ Creation date - 09/14/2018
@ Description - Downloading parcels from TT Prod
'''

import time
import openpyxl
from Utilites import AppConstants
from Utilites.SeleniumOperations import SeleniumOperations
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from Applications.Workflows.ProcessTestFiles.Scripts.Process_Test_Files_Utility import Process_Test_Files_Utility
from Applications.Workflows.ProcessTestFiles.AppResources import CommonLocators


class TT_File_Download:

    def __init__(self, task_type, lo, username):
        self.v_task_type = task_type
        self.v_input_wb = openpyxl.load_workbook(AppConstants.INPUT_FILE_PATH)
        self.lo = lo
        self.v_username = username
        self.v_task_type = task_type
        self.v_data_sheet = self.v_input_wb.get_sheet_by_name("PROCESS_TEST_FILES_INPUT")


    #Method for Search Supplier, Retailer,Doctype and getting parcels
    def process(self, path):
        self.lo.log_to_file("INFO", "Executing TT_File_Download.process()")
        supplier_name = self.v_input_sheet.cell(row=2, column=1).value
        retailer_name = self.v_input_sheet.cell(row=2, column=2).value
        doc_type = self.v_input_sheet.cell(row=2, column=3).value
        date = self.v_input_sheet.cell(row=2, column=4).value
        ptfu = Process_Test_Files_Utility(self.v_task_type, self.v_driver, self.lo)
        so = SeleniumOperations(self.v_task_type,self.v_driver,self.lo)
        time.sleep(3)
        self.v_driver.switch_to.frame(0)
        ptfu.search_by_names(supplier_name,retailer_name,doc_type,date)
        ptfu.get_five_parcels()
        path = CommonLocators.PROCESS_PATH

        #After Getting parcels, search with parcelID
        with open(path) as f:
            for line in f:
                parcel=line.replace('\n', '')
                ptfu.search_by_parcel_id(str(parcel))
                so.click_element_by_xpath(CommonLocators.TRANSAFORMATION)
                element = WebDriverWait(self.v_driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, CommonLocators.TRANSAFORMATION1)))
                element.click()
                time.sleep(2)
                if 'str' in line:
                    break


    def execute_main(self):
        self.lo.log_to_file("INFO", "Executing TT_File_Download.execute_main()")
        time.sleep(5)
        self.process(self.v_input_wb)