'''

@ Author - Ayushi Pawar
@ Creation date - 09/14/2018
@ Description - For uploading file on FTP
'''

import time
import os
from Utilites.SeleniumOperations import SeleniumOperations, AppConstants, LogFileUtility
from Applications.Workflows.ProcessTestFiles.AppResources import CommonLocators
import openpyxl
from easygui import *
import getpass

class File_upload_autoit:
    def __init__(self, task_type, lo, username,driver):
        self.v_input_wb = openpyxl.load_workbook(AppConstants.INPUT_FILE_PATH)
        self.v_task_type = task_type
        self.v_driver = driver
        self.v_input_sheet = self.v_input_wb.get_sheet_by_name('PROCESS_TEST_FILES_INPUT')
        self.lo = lo
        self.v_username = username
        print("in init")

    #Remove_delimiters like tilt
    def file_len(self,fname):
        with open(fname) as f:
            line_count = 0
            for line in f:
                line_count += 1
        return line_count
    def remove_delimiters(self,fname):
        self.lo.log_to_file("INFO", "Executing File_upload_autoit.remove_delimiters()")
        path="C://Users//"+getpass.getuser()+"//Downloads//"+fname+".dat"
        file_lens = self.file_len(path)
        file1 = open(path, "r")
        data = file1.read()
        # if only one line present in file
        if file_lens == 1:
            for ch in ['~', '|', '^', '☐']:
                if ch in data:
                    data = data.replace(ch, "\n")
        else:
            for ch in ['~', '|', '^', '☐']:
                if ch in data:
                    data = data.replace(ch, "")
        file1.close()
        file1 = open(path, "w")
        file1.write(data)
        file1.close()

    #validating ISA IDs
    def validation(self,parcel, Excel_filepath):
        self.lo.log_to_file("INFO", "Executing File_upload_autoit.validation()")
        path = "C://Users/" + getpass.getuser() + "/Downloads/" + parcel + ".dat"
        path1=open(path, 'r')

        def search_excel_file(Excel_filepath, row, column):
            wb_obj = openpyxl.load_workbook(Excel_filepath)
            sheet_obj = wb_obj.active
            cell_obj = sheet_obj.cell(row=row, column=column)
            return str(cell_obj.value)

        def search_EDI_file(path1):
            for line in path1:
                fields = line.split('*')
                if fields[0] == "ISA":
                    return str(str(fields[6]) + "-" + str(len(str(fields[6]))) + "-" + str(fields[8]) + "-" + str(
                        len(str(fields[8]))))
                    break

        ISA_IDS = search_EDI_file(path1)
        array1 = ISA_IDS.split("-")
        edi_retailer_isa_id = (str(array1[0])).strip()
        edi_retailer_isa_id_length = array1[1].strip()
        edi_supplier_isa_id = str(array1[2]).strip()
        edi_supplier_isa_id_length = array1[3].strip()
        excel_retailer_isa_id = str(search_excel_file(Excel_filepath, 2, 8).strip())
        excel_supplier_isa_id = str(search_excel_file(Excel_filepath, 2, 6).strip())
        print(edi_retailer_isa_id)
        print(excel_retailer_isa_id)
        if edi_retailer_isa_id == excel_retailer_isa_id:
            msg1 = "RETAILER: FILE ISA ID & DC4 ISA ID are matching."
        else:
            msg1 = "RETAILER: FILE ISA ID & DC4 ISA ID are NOT matching."
        if edi_supplier_isa_id == excel_supplier_isa_id:
            msg2 = "SUPPLIER: FILE ISA ID & DC4 ISA ID are matching."
        else:
            msg2 = "SUPPLIER: FILE ISA ID & DC4 ISA ID are NOT matching."
        msg = msg1 + "\n\n" + msg2
        title = "ISA ID Validation"
        if ccbox(msg, title):
            flag=1
            return flag
            pass
        else:
            flag=0
            return flag


    #Upload Files on FTP
    def upload_file(self,driver,parcel):
        self.lo.log_to_file("INFO", "Executing File_upload_autoit.upload_file()")
        so = SeleniumOperations(self.v_task_type, self.v_driver, self.lo)
        time.sleep(3)
        so.click_element_by_xpath(AppConstants.FTP_ADD_FILES_BTN)
        time.sleep(2)
        #Hardcoded path required for AutoIT tool
        path="C:" + "\\" + "Users" + "\\" + getpass.getuser() + "\\" + "Downloads" + "\\" +parcel + ".dat"
        os.system(AppConstants.AUTOIT_FTP_UPLOAD_SCRIPT + path)
        time.sleep(5)
        driver.switch_to.frame(2)
        so.click_element_by_xpath(C.FTP_UPLOAD_FILES_BTN)
        time.sleep(2)
        i=1
        for i in range (2):
            driver.refresh()
            time.sleep(10)

    def execute_main(self):
        self.lo.log_to_file("INFO", "Executing File_upload_autoit.execute_main()")
