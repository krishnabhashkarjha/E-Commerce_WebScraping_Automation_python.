'''

@ Author - Aditya Datar
@ Creation date - 09/14/2018
@ Description - Main Script of Review Setup
'''
import openpyxl
import time
from Utilites.Login import Login
from Utilites import AppConstants
from selenium import webdriver
from Utilites.SeleniumOperations import SeleniumOperations
from Applications.Workflows.ProcessTestFiles.AppResources import CommonLocators
from Applications.Workflows.ProcessTestFiles.Scripts.Process_Test_File_850 import Process_Test_file_850


class Process_Test_Files:
    def __init__(self, task_type, lo, username):
        self.v_task_type = task_type
        self.v_driver = webdriver.Chrome(AppConstants.BROWSER_DRIVER)
        self.v_input_wb = openpyxl.load_workbook(AppConstants.INPUT_FILE_PATH)
        self.v_data_sheet = self.v_input_wb.get_sheet_by_name('PROCESS_TEST_FILES_INPUT')
        self.lo = lo
        self.v_username = username

    def execute_main(self):
        self.lo.log_to_file("INFO", "Executing Process_Test_Files.execute_main()")
        lg = Login(self.v_task_type, self.v_driver, self.v_input_wb, self.lo)
        so = SeleniumOperations(self.v_task_type, self.v_driver, self.lo)
        self.v_data_sheet = self.v_input_wb.get_sheet_by_name('PROCESS_TEST_FILES_INPUT')
        total_rows = self.v_data_sheet.max_row
        lg.login("Launchpad")
        self.v_driver.maximize_window()
        self.v_driver.execute_script("window.open('about:blank', 'tab2');")
        self.v_driver.switch_to.window("tab2")
        lg.login("DC4 PreProd")
        self.v_driver.execute_script("window.open('about:blank', 'tab3');")
        self.v_driver.switch_to.window("tab3")
        lg.login("FTP Pre-prod")
        for row_count in range(2,total_rows+1):
            v_supplier = self.v_data_sheet.cell(row=row_count, column=1).value
            v_retailer = self.v_data_sheet.cell(row=row_count, column=2).value
            v_doc_type = self.v_data_sheet.cell(row=row_count, column=3).value
            v_date = self.v_data_sheet.cell(row=row_count, column=4).value
            if v_doc_type == CommonLocators.PO_File:
                ptf_850 = Process_Test_file_850(self.v_task_type,self.lo,self.v_username,self.v_input_wb, self.v_driver)
                ptf_850.execute_main(v_supplier, v_retailer, v_doc_type,row_count,v_date)
            else:
                self.lo.log_to_file("ERROR", "Document is not found in Process_Test_Files.execute_main()")
