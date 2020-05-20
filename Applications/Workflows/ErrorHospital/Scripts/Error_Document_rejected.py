'''

@ Author - Karan Pandya
@ Creation date - 08/29/2018
@ Description - Procedure for document rejected error title
'''

import time
import math
import datetime
import openpyxl
from Utilites.Login import Login
from openpyxl.styles import Color, PatternFill, Font, Border
from Utilites import AppConstants
from selenium import webdriver
from Utilites.LogFileUtility import LogFileUtility
from Utilites.DC4_Utility import DC4_Utility
from Utilites.SeleniumOperations import SeleniumOperations
from Utilites.ReportFileUtility import ReportFileUtility
from Applications.Workflows.ErrorHospital.Scripts.ErrorHospital_Utility import ErrorHospital_Utility
from Utilites.ExcelOperations import ExcelOperations
from Applications.Workflows.ErrorHospital.AppResources import LocalElementLocator
from openpyxl import load_workbook
from Applications.Workflows.Requeue.Scripts.Requeue import Requeue
import re

class Error_Document_rejected:

    def __init__(self, task_type, driver, log_file_object, input_wb, error_hospital_machine_learning_object):
        self.v_task_type = task_type
        self.v_driver = driver
        self.log_file_object = log_file_object
        self.v_input_wb = input_wb
        self.error_hospital_machine_learning_object = error_hospital_machine_learning_object

    def execute_main(self, v_error_title, v_ticket_uid, output_sheet, earlier_str):
        v_count = 0
        selenium_operation_object = SeleniumOperations(self.v_task_type, self.v_driver, self.log_file_object)
        list_duplicate_ticket_uid = []
        output_sheet_excel_operation_object = ExcelOperations(self.v_task_type, output_sheet)
        v_output_sheet_curr_row = 2

        error_hospital_utility_object = ErrorHospital_Utility(self.v_task_type, self.v_driver, self.log_file_object, self.error_hospital_machine_learning_object)
        #Search for error tickets
        error_hospital_utility_object.document_rejected_error_search(v_ticket_uid, earlier_str, ' ', v_error_title, ' ')
        if selenium_operation_object.check_exists_by_xpath(LocalElementLocator.EH_SHOW_ALL_TAB_XPATH):
            selenium_operation_object.click_element_by_xpath(LocalElementLocator.EH_SHOW_ALL_TAB_XPATH)
        v_tuid_index = 2
        while (error_hospital_utility_object.get_ticket_uid(v_tuid_index)):
            v_temp_ticket_uid = error_hospital_utility_object.get_ticket_uid(v_tuid_index)
            v_temp_receiver = error_hospital_utility_object.get_receiver_name(v_tuid_index)
            v_temp_sender = error_hospital_utility_object.get_sender_name(v_tuid_index)
            #Search for new unique ticket
            if v_temp_ticket_uid not in list_duplicate_ticket_uid:
                v_count =v_count+1
                v_status = ' '
                error_hospital_utility_object.document_rejected_error_search(v_temp_ticket_uid, earlier_str, ' ', v_error_title, ' ')
                #Get TPID and doctype
                v_tpid, v_doctype = error_hospital_utility_object.get_info_from_description(v_tuid_index - 2)
                #Search for duplicate tickets
                error_hospital_utility_object.document_rejected_error_search(' ', earlier_str, v_tpid, v_error_title, v_doctype)
                if selenium_operation_object.check_exists_by_xpath(LocalElementLocator.EH_SHOW_ALL_TAB_XPATH):
                    selenium_operation_object.click_element_by_xpath(LocalElementLocator.EH_SHOW_ALL_TAB_XPATH)
                v_temp_tuid_index = 2
                parcel_uid_list = []
                #Append all duplicate tickets to list
                while (error_hospital_utility_object.get_ticket_uid(v_temp_tuid_index)):
                    list_duplicate_ticket_uid.append(error_hospital_utility_object.get_ticket_uid(v_temp_tuid_index))
                    parcel_uid_list.append(error_hospital_utility_object.get_parcel_uid(v_temp_tuid_index))
                    doc_type_list = error_hospital_utility_object.get_unique_doc_types(v_doctype)

                    v_temp_tuid_index = v_temp_tuid_index + 1
                #Add info to output sheet
                output_sheet_excel_operation_object.set_value(v_output_sheet_curr_row, 1, v_temp_ticket_uid)
                output_sheet_excel_operation_object.set_value(v_output_sheet_curr_row, 2, v_temp_sender)
                output_sheet_excel_operation_object.set_value(v_output_sheet_curr_row, 3, v_temp_receiver)
                output_sheet_excel_operation_object.set_value(v_output_sheet_curr_row, 4, v_tpid)
                doc_type_string = ','.join(doc_type_list)
                output_sheet_excel_operation_object.set_value(v_output_sheet_curr_row, 5, doc_type_string)
                parcel_uid_string = '\n'.join(parcel_uid_list)
                output_sheet_excel_operation_object.set_value(v_output_sheet_curr_row, 6, parcel_uid_string)

                #Validate TPID
                if len(v_tpid) == 15 and ' ' not in v_tpid:
                    self.log_file_object.log_to_file("INFO", "TPID: "+v_tpid+" is valid")
                    v_status, flag = error_hospital_utility_object.add_trading_partnership(v_error_title, v_tpid, doc_type_list, v_temp_sender, v_temp_receiver, v_status)
                    if flag == 1:
                        first_parcel_list =[]
                        first_parcel_list.append(parcel_uid_list[0])
                        error_hospital_utility_object.requeue(first_parcel_list)
                        time.sleep(2)
                        error_status = error_hospital_utility_object.error_status(parcel_uid_list[0], v_error_title)
                        self.v_driver.switch_to.window(self.v_driver.window_handles[0])
                        time.sleep(2)
                        if error_status == 1:
                            if len(parcel_uid_list) > 1:
                                first_parcel_list = parcel_uid_list[1:]
                                error_hospital_utility_object.requeue(first_parcel_list)
                            v_status = v_status+'\nSuccessfully passed all the parcels'
                            output_sheet_excel_operation_object.set_value(v_output_sheet_curr_row, 7, 'Done')
                            output_sheet.cell(row=v_output_sheet_curr_row, column=7).fill = PatternFill(start_color='008000',
                                                                                     end_color='008000',
                                                                                     fill_type='solid')
                            self.log_file_object.log_to_file("INFO", "Successfully passed the failed parcels")
                        else:
                            output_sheet_excel_operation_object.set_value(v_output_sheet_curr_row, 7, 'Failed')
                            output_sheet.cell(row=v_output_sheet_curr_row, column=7).fill = PatternFill(
                                start_color='FFFF0000',
                                end_color='FFFF0000',
                                fill_type='solid')
                            self.log_file_object.log_to_file("INFO", "Unsuccessfully passed the failed parcels")

                    else:
                        output_sheet_excel_operation_object.set_value(v_output_sheet_curr_row, 7, 'Push to Manual')
                        output_sheet.cell(row=v_output_sheet_curr_row, column=7).fill = PatternFill(
                            start_color='FFFF0000',
                            end_color='FFFF0000',
                            fill_type='solid')

                    output_sheet_excel_operation_object.set_value(v_output_sheet_curr_row, 8, v_status)

                else:
                    self.log_file_object.log_to_file("INFO", "TPID: " + v_tpid + " is invalid")
                    v_status = v_status + '\nInvalid Data, Report to implementation analyst'
                    output_sheet_excel_operation_object.set_value(v_output_sheet_curr_row, 7, 'Done')
                    output_sheet.cell(row=v_output_sheet_curr_row, column=7).fill = PatternFill(start_color='008000',
                                                                                              end_color='008000',
                                                                                              fill_type='solid')
                    output_sheet_excel_operation_object.set_value(v_output_sheet_curr_row, 8, v_status)

                v_output_sheet_curr_row = v_output_sheet_curr_row + 1
                self.v_input_wb.save(AppConstants.INPUT_FILE_PATH)
                error_hospital_utility_object.document_rejected_error_search(v_ticket_uid, earlier_str, ' ', v_error_title, ' ')
                if selenium_operation_object.check_exists_by_xpath(LocalElementLocator.EH_SHOW_ALL_TAB_XPATH):
                    selenium_operation_object.click_element_by_xpath(LocalElementLocator.EH_SHOW_ALL_TAB_XPATH)
            v_tuid_index = v_tuid_index + 1
        return v_count


