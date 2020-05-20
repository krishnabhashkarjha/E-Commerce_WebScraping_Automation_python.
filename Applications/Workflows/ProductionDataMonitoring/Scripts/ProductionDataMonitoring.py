'''

@ Author - Yogesh Pawar
@ Creation date - 08/29/2018
@ Description - Main Script of Production Data Monitoring
'''
import time
import math
import datetime
import openpyxl
from Utilites.Login import Login
from Utilites import AppConstants
from selenium import webdriver
from Utilites.ReportFileUtility import ReportFileUtility
from Applications.Workflows.ProductionDataMonitoring.AppResources import LocalElementLocator
from Applications.Workflows.ProductionDataMonitoring.Scripts.TransactionTrackerOperations2 import TransactionTrackerOperations
from Utilites.ExcelOperations import ExcelOperations


class ProductionDataMonitoring:

    def __init__(self, task_type, lo, username):
        self.v_input_wb = openpyxl.load_workbook(AppConstants.INPUT_FILE_PATH)
        self.v_task_type = task_type
        self.v_input_sheet = self.v_input_wb.get_sheet_by_name("InputData")
        self.lo = lo
        self.v_username = username

    # method to sort "InProgress" tasks and "Monitoring Done" tasks
    def TaskFilter(self, path, status):
        self.lo.log_to_file("INFO", "Executing ProductionDataMonitoring.TaskFilter()")
        rf = ReportFileUtility(self.v_task_type)
        row_count = self.v_input_sheet.max_row
        if status == 'InProgress':
            self.lo.log_to_file("INFO",
                                "select InProgress task from input sheet from ProductionDataMonitoring.TaskFilter()")
            eo = ExcelOperations(self.v_task_type, self.v_input_sheet)
            for count in range(1, row_count + 1):
                v_start_time = time.time()
                if ((self.v_input_sheet.cell(row=count, column=5).value) == "InProgress"):
                    supplier_name = self.v_input_sheet.cell(row=count, column=1).value
                    retailer_name = self.v_input_sheet.cell(row=count, column=2).value
                    doc_type=self.v_input_sheet.cell(row=count, column=3).value
                    date=self.v_input_sheet.cell(row=count, column=4).value
                    tto = TransactionTrackerOperations(self.v_task_type, self.v_driver, self.lo)
                    tto.search_process_for_PDM(supplier_name,retailer_name,doc_type,date, self.v_input_sheet,count)
                    self.v_input_wb.save(AppConstants.INPUT_FILE_PATH)
                    v_end_time = time.time()
                    rf.update_sheet(self.v_username, 1, math.floor(v_end_time - v_start_time),str(datetime.date.today()))
                    found_parcels= self.v_input_sheet.cell(row=count, column=6).value

                    #logic for change status of task
                    if int(found_parcels)>=4:
                        eo.set_value(count,5,"Monitoring done")
                    self.v_input_wb.save(AppConstants.INPUT_FILE_PATH)



        if status == 'Monitoring done':
            for count in range(1, row_count + 1):
                if ((self.v_input_sheet.cell(row=count, column=5).value) == "Monitoring done"):
                    print(self.v_input_sheet.cell(row=count, column=2).value)

    # main method of CMProductionDataMonitoring
    def execute_main(self):
        self.lo.log_to_file("INFO", "Executing ProductionDataMonitoring.execute_main()")
        self.v_driver = webdriver.Chrome(AppConstants.BROWSER_DRIVER)
        self.v_driver.maximize_window()
        lg = Login(self.v_task_type, self.v_driver, self.v_input_wb, self.lo)
        lg.login("Launchpad")
        self.lo.log_to_file("INFO", "Login in to Launchpad")
        time.sleep(5)
        self.v_driver.get(LocalElementLocator.TRANSACTION_TRACKER_PROD_LINK)
        self.TaskFilter(self.v_input_wb,'InProgress')
        self.lo.log_to_file("INFO", "completed ProductionDataMonitoring.execute_main()")
        self.v_driver.close()

