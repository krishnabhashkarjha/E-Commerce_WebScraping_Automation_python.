# my_list = ['Pune', 'Mumbai', 'Thane']
#
# # from xlwt import Workbook
# # wb = Workbook()
# # sheet1 = wb.add_sheet('Sheet 1')
# # for i in range (0,3):
# #     sheet1.write(0, i, str(my_list[i]))
# #     wb.save('C:/Users/Krishnabhashkar.Jha/Desktop/ListTest.xls')
#
#
# # import xlrd
# # wb = xlrd.open_workbook("C:/Users/Krishnabhashkar.Jha/Desktop/ListTest.xls")
# # sheet = wb.sheet_by_index(0)
# # for i in range(0,3):
# #     # print(i)
# #     print(sheet.cell(0,i).value)
#
#
#
#
# # import openpyxl
# # wb = openpyxl.load_workbook("C:/Users/Krishnabhashkar.Jha/Desktop/ListTest2.xlsx")
# # sheet = wb.get_sheet_by_name("Sheet1")
# # print(sheet.cell(row=1,column=1).value)
#
#
#
# import smtplib, ssl
#
# port = 465  # For SSL
# smtp_server = "smtp.gmail.com"
# sender_email = "bhashkar.jha17@gmail.com"  # Enter your address
# receiver_email = "abhishek.wazalwar@gmail.com"  # Enter receiver address
# password = input("Type your password and press enter: ")
# message = """\
# Subject: Hi there
#
# This message is sent from Python."""
# try:
#     context = ssl.create_default_context()
#     with smtplib.SMTP_SSL(smtp_server, port, context=context) as server:
#         server.login(sender_email, password)
#         server.sendmail(sender_email, receiver_email, message)
# except Exception as e:
#     print(e)
#
# arr = [1,2,3,4,5]
# b = len(arr);print(b)
# err = [6,7,8,9,0]
# c = len(err);print(c)
#
# for i in arr:
#     for j in err:
#         d = i+j
#         # print(d)
#
# sum = 0
# for i in arr:
#     sum = sum+i
# print(sum)

# for i in arr:
#     if i>2:
#         print(i)

import django
import openpyxl


def get_value(row_index, column_index):
    print("1")
    file_name = r"C:\Users\Krishnabhashkar.Jha\Desktop\InputSheet.xlsx"
    wb = openpyxl.load_workbook(file_name)
    print("2")
    sheet = wb.active
    v_value = sheet.cell(row=row_index, column=column_index)
    # print(v_value)
    return v_value.value



a = get_value(1,3)
print(a)









