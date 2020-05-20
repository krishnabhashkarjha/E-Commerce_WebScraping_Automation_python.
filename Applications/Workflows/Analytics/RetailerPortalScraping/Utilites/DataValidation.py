"""
@ Author - Krishnabhashkar jha
@ Creation date - 13/02/2020
@ Description - All Data validation code.
"""

import os
import csv
import getpass
import glob
import openpyxl
from functools import reduce
import math
from Applications.Workflows.Analytics.RetailerPortalScraping.Utilites.LogFileUtility import LogFileUtility as lo

class DataValidation:

    def DataValidation(self):
        try:
            v_input_wb = openpyxl.load_workbook("C:/Users/Krishnabhashkar.Jha/Desktop/Book1.xlsx")
            v_input_sheet = v_input_wb.get_sheet_by_name("Sheet2")
            total_rows_input = v_input_sheet.max_row
            arr = []
            for i in range(2, int(total_rows_input) + 1):
                j = 5
                while True:
                    arr.append(v_input_sheet.cell(row=(i), column=j).value)
                    j += 1
                    if type(v_input_sheet.cell(row=(i), column=j).value) != int:
                        break
                print(arr)
                sum = reduce(lambda a, b: a + b, arr)
                w_bar = int(sum) / int(len(arr))
                w_bar_arr = list(map(lambda w: w - w_bar, arr))
                square_arr = list(map(lambda g: g * g, w_bar_arr))
                addition_arr = reduce(lambda k, l: k + l, square_arr)
                sd = addition_arr / (int(len(arr)) - 1)
                Final_SD = math.sqrt(sd)
                upper_limit = (w_bar + (3 * Final_SD))
                lower_limit = (w_bar - (3 * Final_SD))
                # Final_SD = math.sqrt(addi_arr)
                # print(w_bar_arr)
                # print(square_arr)
                # print("w_bar = "+str(w_bar))
                # print("SD is : "+str(sd))
                print("Final SD is : " + str(Final_SD))
                print("Upper Limit of File is : " + str(upper_limit))
                print("Lower Limit of File is : " + str(lower_limit))
                v_input_sheet.cell(row=i, column=2).value = lower_limit
                v_input_sheet.cell(row=i, column=3).value = upper_limit
                v_input_sheet.cell(row=i, column=4).value = Final_SD
                v_input_wb.save("C:/Users/Krishnabhashkar.Jha/Desktop/Book1.xlsx")#Give file name here
                arr.clear()
                w_bar_arr.clear()
                square_arr.clear()
        except Exception as e:
            lo.log_to_file(self, "ERROR", "Exception :" + str(e))

    def FileValidation(self):
        try:
            list_of_files = glob.glob("C:\\Users\\" + getpass.getuser() + "\\Downloads\\*")
            latest_file_In_Download = max(list_of_files, key=os.path.getctime)
            print(latest_file_In_Download)
            with open(latest_file_In_Download) as latest_file_In_Download:
                wb = csv.reader(latest_file_In_Download)
                i = 1
                for row in wb:
                    if i == 2:
                        a = row[12]
                        break
                    i += 1
                if a == "BPS":
                    print("This is BPS File")
                elif a == "CAB":
                    print("This is CAB File")
                else:
                    print("No file is there")
        except Exception as e:
            lo.log_to_file(self,"ERROR", "Exception : "+str(e))


