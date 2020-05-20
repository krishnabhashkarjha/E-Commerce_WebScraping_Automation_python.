'''

@ Author - Karan Pandya
@ Creation date - 10/19/2018
@ Description - Main Script of Web Fulfillment
'''
import time
import math
import datetime
import openpyxl
from Utilites.Login import Login
from Utilites import AppConstants
from selenium import webdriver
from Utilites.LogFileUtility import LogFileUtility
from Utilites.DC4_Utility import DC4_Utility
from Utilites.SeleniumOperations import SeleniumOperations
from Utilites.ReportFileUtility import ReportFileUtility
from Applications.Workflows.ErrorHospital.Scripts.ErrorHospital_Utility import ErrorHospital_Utility
from Utilites.ExcelOperations import ExcelOperations
from Applications.Workflows.Covalent_Web_Fulfillment.AppResources import LocalElementLocator
from openpyxl import load_workbook
from Applications.Workflows.ErrorHospital.Scripts.Error_Document_rejected import Error_Document_rejected
import re
from Applications.Workflows.Covalent_Web_Fulfillment.Scripts.Web_Fulfillment_Utility import Web_Fulfillment_Utility
from Utilites.DatabaseUtility import DatabaseUtility
from Utilites.ReportFileUtility import ReportFileUtility
import easygui
import threading

class Covalent_Web_Fulfillment_main:
    def __init__(self, task_type, lo, username):
        self.v_task_type = task_type
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_argument("--incognito")
        chrome_options.add_argument("--start-maximized")
        chrome_options.add_argument("--window-size=900,900")
        self.v_driver_mv = webdriver.Chrome(AppConstants.BROWSER_DRIVER, chrome_options=chrome_options)
        self.v_driver = webdriver.Chrome(AppConstants.BROWSER_DRIVER, chrome_options= chrome_options)
        self.v_input_wb = openpyxl.load_workbook(AppConstants.INPUT_FILE_PATH)
        self.lo = lo
        self.v_username = username
        #self.v_output_wb = openpyxl.load_workbook("../X-Runner/Output.xlsx")
    def execute_main(self):


        t1 = threading.Thread(target=self.thread1)

        t2 = threading.Thread(target=self.thread2)
        t1.start()
        t2.start()

        t1.join()
        t2.join()

        db_object = DatabaseUtility(self.v_task_type, self.lo)
        if db_object.connection():
            self.lo.log_to_file('INFO', 'Successful Connection')
        else:
            self.lo.log_to_file('ERROR', 'DB Server is not active')
            return
        v_start_time = time.time()
        self.lo.log_to_file("INFO", "Login in to DC4 Prod")
        login_obj = Login(self.v_task_type, self.v_driver, self.v_input_wb, self.lo)
        selenium_operation_obj = SeleniumOperations(self.v_task_type, self.v_driver, self.lo)
        report_file_obj = ReportFileUtility(self.v_task_type)
        input_sheet = self.v_input_wb.get_sheet_by_name('Input')
        input_sheet_row_count = input_sheet.max_row
        excel_operation_obj = ExcelOperations(self.v_task_type, input_sheet)
        Web_Fulfillment_Utility_obj = Web_Fulfillment_Utility(self.v_task_type, self.v_driver, self.v_input_wb, self.lo, self.v_username, db_object, self.v_driver_mv)
        # login_obj.login("Salesforce")
        self.v_driver.execute_script(
            "window.open('http://dc4ui.p01.pro:7777/dc4custmanager/faces/home.jsp', 'new_window')")


        # self.v_driver.execute_script(
        #     "window.open('https://atlassian.spscommerce.com/browse/FICS-163692', 'new_window')")
        self.v_driver.switch_to.window(self.v_driver.window_handles[-1])
        report_file_object = ReportFileUtility(self.v_task_type)
        time.sleep(2)
        login_obj.login("DC4 Prod")

        input_row = input_sheet_row_count + 1
        FICS_Number = excel_operation_obj.get_value(input_row, 1)
        flag = 1
        count  = 0
        v_start_time = time.time()
        while(flag):
            v_start_time1 = time.time()
            v_salesforce_id = easygui.enterbox('Please enter salesforce ID')
            self.v_driver.switch_to.window(self.v_driver.window_handles[0])
            time.sleep(2)
            #self.v_driver.get('https://atlassian.spscommerce.com/browse/'+FICS_Number)
            #v_supplier_name, v_email_address, v_phone_number, v_retailer_name, v_billing_address, v_vendor_number, v_retailer_web_company_uid, v_account_number, v_retailer_dc4_company_uid, v_contact_name, v_salesforce_id = Web_Fulfillment_Utility_obj.get_info_from_JIRA()
            v_supplier_name, v_email_address, v_phone_number, v_retailer_name, v_billing_address, v_vendor_number, v_retailer_web_company_uid, v_account_number, v_retailer_dc4_company_uid, v_contact_name, client_uid, retailer_uid = Web_Fulfillment_Utility_obj.get_info_from_salesforce(v_salesforce_id)
            print(
                 v_salesforce_id + "\n" + v_supplier_name + "\n" + v_email_address + "\n" + v_phone_number + "\n" + v_retailer_name + "\n" + v_vendor_number + "\n" + v_retailer_web_company_uid + "\n" + v_account_number + "\n" + v_retailer_dc4_company_uid + "\n" + v_contact_name)
            print(v_billing_address)

            self.v_driver.switch_to.window(self.v_driver.window_handles[-1])
            time.sleep(2)
            #Web_Fulfillment_Utility_obj.get_data_for_library(v_salesforce_id, v_supplier_name, v_retailer_name, v_retailer_dc4_company_uid, FICS_Number)
            Web_Fulfillment_Utility_obj.run_setup_tool(v_salesforce_id, v_supplier_name, v_email_address, v_phone_number, v_retailer_name, v_billing_address, v_vendor_number, v_retailer_web_company_uid, v_account_number, v_retailer_dc4_company_uid, v_contact_name, client_uid, retailer_uid)
            print(input_row)
            count = count+1
            excel_operation_obj.set_value(input_row,1,count)




            FICS_Number = excel_operation_obj.get_value(input_row, 1)
            v_end_time1 = time.time()
            excel_operation_obj.set_value(input_row, 2, math.floor(v_end_time1 - v_start_time1))
            excel_operation_obj.set_value(input_row, 3, datetime.date.today())
            print('----------------------------------------------\n\nExecution Time: '+str(math.floor(v_end_time1 - v_start_time1))+'   count: '+str(count)+'\n\n-------------------------')
            time.sleep(2)
            input_row = input_row + 1

            self.v_input_wb.save(AppConstants.INPUT_FILE_PATH)
            flag = easygui.ynbox('Do you want to run one more setup?')
        v_end_time = time.time()
        report_file_object.update_sheet(self.v_username, count, math.floor(v_end_time - v_start_time),
                                        str(datetime.date.today()))

    def thread1(self,):
        login_obj = Login(self.v_task_type, self.v_driver, self.v_input_wb, self.lo)
        login_obj.login("Salesforce")
        time.sleep(5)
    def thread2(self):
        login_obj = Login(self.v_task_type, self.v_driver_mv, self.v_input_wb, self.lo)
        login_obj.login('Launchpad')
        time.sleep(5)
