import gspread
from oauth2client.service_account import ServiceAccountCredentials
from Utilites import AppConstants
import datetime
import openpyxl
class Web_Fulfillment_machine_learning:

    def __init__(self, task_type, username, v_output_wb):
        self.v_task_type = task_type
        # scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        # creds = ServiceAccountCredentials.from_json_keyfile_name(AppConstants.CLIENT_SECRET_JSON_FILE_PATH, scope)
        # client = gspread.authorize(creds)
        # sht = client.open("Web Fulfillment")
        # self.G_supplier_sheet = sht.worksheet('Supplier Info')
        # self.G_retailer_sheet = sht.worksheet('Retailer Info')
        # values_list = self.G_supplier_sheet.col_values(1)
        # Row_Count = len(values_list)
        # self.supplier_row_count = Row_Count + 1
        # values_list = self.G_retailer_sheet.col_values(1)
        # Row_Count = len(values_list)
        # self.retailer_row_count = Row_Count + 1
        #
        self.v_username = username
        self.v_output_wb = v_output_wb
        self.supplier_sheet = self.v_output_wb.worksheets[0]
        self.retailer_sheet = self.v_output_wb.worksheets[1]
        self.retailer_row_count = self.retailer_sheet.max_row
        self.supplier_row_count = self.supplier_sheet.max_row
        # print(str(self.supplier_row_count))
        # print(str(self.retailer_row_count))


    # def find_supplier_dc4_company_in_library(self, company_name):
    #     for i in range(1, self.supplier_row_count):
    #         if self.G_supplier_sheet.cell(i,1).value == company_name:
    #             return self.G_supplier_sheet.cell(i,3).value
    #
    #     return False

    # def add_supplier_data_to_library(self, jira_supplier_name, dc4_supplier_name, supplier_dc4_id, supplier_web_company_uid, supplier_web_member_uid, fics_number):
    #     flag = 0
    #     for i in range(1, self.supplier_row_count):
    #         if self.G_supplier_sheet.cell(i, 1).value == jira_supplier_name:
    #             flag = 1
    #             break
    #     if flag == 1:
    #         return
    #     else:
    #         val = []
    #         val.append(jira_supplier_name)
    #         val.append(dc4_supplier_name)
    #         val.append(supplier_dc4_id)
    #         val.append(supplier_web_company_uid)
    #         val.append(supplier_web_member_uid)
    #         val.append(self.v_username)
    #         val.append(str(datetime.date.today()))
    #         val.append(fics_number)
    #         self.G_supplier_sheet.append_row(val)

    # def add_retailer_data_to_library(self, jira_retailer_name, dc4_retailer_name, retailer_dc4_id, retailer_web_company_uid, TPID, fics_number):
    #     flag = 0
    #     for i in range(1, self.retailer_row_count):
    #         if self.G_retailer_sheet.cell(i, 1).value == jira_retailer_name:
    #             flag = 1
    #             break
    #     if flag == 1:
    #         return
    #     else:
    #         val = []
    #         val.append(jira_retailer_name)
    #         val.append(dc4_retailer_name)
    #         val.append(retailer_dc4_id)
    #         val.append(retailer_web_company_uid)
    #         for i in range(0, len(TPID)):
    #             val.append(TPID[i])
    #         self.G_retailer_sheet.append_row(val)
    #         self.G_retailer_sheet.update_cell(self.retailer_row_count, 14, self.v_username)
    #         self.G_retailer_sheet.update_cell(self.retailer_row_count, 15, str(datetime.date.today()))
    #         self.G_retailer_sheet.update_cell(self.retailer_row_count, 16, fics_number)

    def add_supplier_data_to_library(self, jira_supplier_name, dc4_supplier_name, supplier_dc4_id, supplier_web_company_uid, supplier_web_member_uid, fics_number):
        flag = 0

        for i in range(1, self.supplier_row_count+1):
            if self.supplier_sheet.cell(i, 1).value == jira_supplier_name:
                flag = 1
                break
        if flag == 1:
            return
        elif flag == 0 and dc4_supplier_name != '':
            self.supplier_sheet.cell(self.supplier_row_count + 1, 1).value = jira_supplier_name
            self.supplier_sheet.cell(self.supplier_row_count + 1, 2).value = dc4_supplier_name
            self.supplier_sheet.cell(self.supplier_row_count + 1, 3).value = supplier_dc4_id
            self.supplier_sheet.cell(self.supplier_row_count + 1, 4).value = supplier_web_company_uid
            self.supplier_sheet.cell(self.supplier_row_count + 1, 5).value = supplier_web_member_uid
            self.supplier_sheet.cell(self.supplier_row_count + 1, 6).value = self.v_username
            self.supplier_sheet.cell(self.supplier_row_count + 1, 7).value = str(datetime.date.today())
            self.supplier_sheet.cell(self.supplier_row_count + 1, 8).value = fics_number

            self.v_output_wb.save('../X-Runner/Output.xlsx')


    def add_retailer_data_to_library(self, jira_retailer_name, dc4_retailer_name, retailer_dc4_id, retailer_web_company_uid, TPID, fics_number):
        flag = 0
        for i in range (1, self.retailer_row_count+1):
            if self.retailer_sheet.cell(i, 1).value == jira_retailer_name:
                flag = 1
                break
        if flag == 1:
            return
        elif flag == 0 and dc4_retailer_name != '':
            self.retailer_sheet.cell(self.retailer_row_count + 1, 1).value = jira_retailer_name
            self.retailer_sheet.cell(self.retailer_row_count + 1, 2).value = dc4_retailer_name
            self.retailer_sheet.cell(self.retailer_row_count + 1, 3).value = retailer_dc4_id
            self.retailer_sheet.cell(self.retailer_row_count + 1, 4).value = retailer_web_company_uid
            for i in range(0, len(TPID)):
                self.retailer_sheet.cell(self.retailer_row_count + 1, 5+i).value = TPID[i]

            self.retailer_sheet.cell(self.retailer_row_count, 14).value =  self.v_username
            self.retailer_sheet.cell(self.retailer_row_count, 15).value =  str(datetime.date.today())
            self.retailer_sheet.cell(self.retailer_row_count, 16).value =  fics_number
            self.v_output_wb.save('../X-Runner/Output.xlsx')

