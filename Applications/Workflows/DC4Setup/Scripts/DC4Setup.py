#<Description>
#<Author>
#<Creation Date>
#<Updated By>


import openpyxl
import time
import math
import datetime
from selenium import webdriver
from Utilites.LogFileUtility import LogFileUtility
#from Utilites.Execution import Execution
from Utilites import AppConstants
from Utilites.Login import Login
from Utilites.SeleniumOperations import SeleniumOperations
from Utilites.ReportFileUtility import ReportFileUtility
from Utilites.DC4_Utility import DC4_Utility
from Applications.Workflows.DC4Setup.AppResources.DC4_Utility_Local import DC4_Utility_Local
import easygui

class DC4Setup:
    def __init__(self, task_type, lo, username):
        self.v_task_type = task_type
        self.v_driver = webdriver.Chrome(AppConstants.BROWSER_DRIVER)
        self.v_input_wb = openpyxl.load_workbook(AppConstants.INPUT_FILE_PATH)
        self.lo = lo
        self.v_username = username


    def execute_main(self):
            v_start_time = time.time()
            self.lo.log_to_file("INFO", "Login in to DC4 Pre-Prod")
            lg = Login(self.v_task_type, self.v_driver, self.v_input_wb, self.lo)
            so = SeleniumOperations(self.v_task_type, self.v_driver, self.lo)
            input_sheet = self.v_input_wb.get_sheet_by_name('Input')

            #HardCoded Location
            current_row = 2
            Company_Name = input_sheet.cell(row=current_row, column=1).value
            Retailer_Name=input_sheet.cell(row=current_row,column=2).value
            Capability_UID=input_sheet.cell(row=current_row,column=5).value

            #Login to DC4 Pre-Production
            lg.login("DC4 PreProd")

            #Entering & Searching DC4 Company
            dc4_common_dc4_utility_object = DC4_Utility(self.v_task_type, self.v_driver, self.lo)
            dc4_common_dc4_utility_object.company_search_by_name(Company_Name)

            #Clicking on Searched Company Name, Relationships Tab
            dc4_local_dc4_utility_local_object=DC4_Utility_Local(self.v_task_type, self.v_driver, self.lo, self.v_input_wb)
            dc4_local_dc4_utility_local_object.click_on_searched_company_name()
            dc4_local_dc4_utility_local_object.add_new_capability()
            dc4_local_dc4_utility_local_object.return_capability_uid()
            dc4_local_dc4_utility_local_object.click_on_preprod_relationships_tab()
            dc4_local_dc4_utility_local_object.search_trading_partner_name_in_relationships(Retailer_Name)
            dc4_local_dc4_utility_local_object.click_on_profile_name_via_relationships()
            dc4_local_dc4_utility_local_object.click_on_show_link_of_searched_profile()
            dc4_local_dc4_utility_local_object.add_existing_capability()

            time.sleep(10)







            #self.v_driver.close()
            #v_end_time = time.time()
            #rf.update_sheet(self.v_username, 2, math.floor(v_end_time - v_start_time), str(datetime.date.today()))
