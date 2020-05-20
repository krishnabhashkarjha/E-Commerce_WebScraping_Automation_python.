from Utilites.Login import Login
from selenium import webdriver
from Utilites import AppConstants
import openpyxl
import xlrd
from Utilites.SeleniumOperations import SeleniumOperations
from Utilites.ExcelOperations import ExcelOperations
import time
from Applications.Workflows.QB_Setup_Process.Scripts.QB_Setup_Utility import QB_Setup_Util



class QB_Setup:
    def __init__(self, task_type, lo, username):
        self.v_task_type = task_type
        self.v_driver = webdriver.Chrome(AppConstants.BROWSER_DRIVER)
        self.v_input_wb = openpyxl.load_workbook(AppConstants.INPUT_FILE_PATH)
        self.v_data_sheet = self.v_input_wb.get_sheet_by_name('QB_Setup')
        self.lo = lo
        self.v_username = username


    def execute_main(self):
        self.lo.log_to_file("INFO", "Login in to DC4 Prod")
        login_object = Login(self.v_task_type, self.v_driver, self.v_input_wb, self.lo)
        selenium_operation_object = SeleniumOperations(self.v_task_type, self.v_driver, self.lo)
        # report_file_object = ReportFileUtility(self.v_task_type)

        total_rows = self.v_data_sheet.max_row
        print(total_rows)
        login_object.login("DC4 Prod")
        self.v_driver.maximize_window()

        for row_count in range(2,total_rows+1):
            v_start_time = time.time()
            excel_operations = ExcelOperations(self.v_task_type, self.v_data_sheet)
            v_supplier = self.v_data_sheet.cell(row=row_count, column=1).value
            v_retailer = self.v_data_sheet.cell(row=row_count, column=2).value
            v_erp = self.v_data_sheet.cell(row=row_count,column=15).value
            print(v_retailer)
            print(v_erp)
            service_name = ["FItoService", "FIfromService"]

            v_doc_type = self.v_data_sheet.cell(row=row_count, column=3).value
            print(v_doc_type)
            v_date = self.v_data_sheet.cell(row=row_count, column=4).value

            if v_doc_type == 850:
                print(v_doc_type)
                Setup_U_Ob = QB_Setup_Util(self.v_task_type,self.lo,self.v_username,self.v_input_wb,self.v_driver)
                retailer_flag = Setup_U_Ob.retailer_v_check(v_supplier, v_retailer, v_doc_type, row_count)
                supplier_flag = Setup_U_Ob.Supplier_Setup_Check(v_supplier,v_retailer,v_doc_type, v_erp, row_count)






