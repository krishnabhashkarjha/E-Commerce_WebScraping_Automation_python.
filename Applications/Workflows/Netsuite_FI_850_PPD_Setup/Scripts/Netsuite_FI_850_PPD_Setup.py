'''

@ Author - Purva Tapre, Aditee Avasarikar
@ Creation date - 09/20/2018
@ Description - Main Script of Netsuite_FI_850_PPD_Setup
'''
import time
import math
import datetime
import openpyxl

from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import Select
from Utilites.LogFileUtility import LogFileUtility

from Utilites.Login import Login
from Utilites import AppConstants
from selenium import webdriver
from Utilites.ExcelOperations import ExcelOperations
from Utilites.LogFileUtility import LogFileUtility
from Utilites.DC4_Utility import DC4_Utility
from Utilites.SeleniumOperations import SeleniumOperations
from Utilites.ReportFileUtility import ReportFileUtility
from Applications.Workflows.Netsuite_FI_850_PPD_Setup.AppResources import LocalElementLocator

class Netsuite_FI_850_PPD_Setup:
    def __init__(self, task_type, lo, username):
        self.v_task_type = task_type
        self.v_driver = webdriver.Chrome(AppConstants.BROWSER_DRIVER)
        self.v_input_wb = openpyxl.load_workbook(AppConstants.INPUT_FILE_PATH)
        self.lo = lo
        self.v_username = username

    def show_all(self,V_path,V_option):
        try:
            dropdown_list = Select(self.v_driver.find_element_by_xpath(V_path))
            dropdown_list.select_by_value(V_option)
        except NoSuchElementException:
            return False
        return True

    def search_retailer_profile(self, profie_name):
        so = SeleniumOperations(self.v_task_type, self.v_driver, self.lo)
        self.show_all('//*[@id="form1:table1-nb__xc_c"]','all')
        profile_flag = 0
        row_index = 2

        while(profile_flag == 0):
            V_relationship_profile_path = '//*[@id="form1:table1"]/table[2]/tbody/tr[' + str(row_index) + ']/td[4]'
            if so.check_exists_by_xpath(V_relationship_profile_path):
                V_relationship_profile_name = so.get_text_by_xpath(V_relationship_profile_path)
                if V_relationship_profile_name == profie_name:
                    ret_row_index = row_index - 2
                    V_relationship_retailer_name_path = '//*[@id="form1:table1:' + str(ret_row_index) + ':commandLink2"]'
                    V_relationship_retailer_profile_path = '//*[@id="form1:table1"]/table[2]/tbody/tr[' + str(row_index) + ']/td[6]'
                    if so.check_exists_by_xpath(V_relationship_retailer_name_path):
                        V_relationship_retailer_name = so.get_text_by_xpath(V_relationship_retailer_name_path)
                        ret_row_index = row_index - 2
                        so.click_element_by_xpath(V_relationship_retailer_name_path)
                        so.click_element_by_xpath('//*[@id="form1:panelPage1"]/table/tbody/tr[3]/td/table/tbody/tr/td[2]/table/tbody/tr/td[3]/a')
                        self.lo.log_to_file("INFO", "Retailer company Found")
                    else:
                        return False
                    if so.check_exists_by_xpath(V_relationship_retailer_profile_path):
                        V_relationship_retailer_profile_name = so.get_text_by_xpath(V_relationship_retailer_profile_path)
                        self.lo.log_to_file("INFO", "Retailer profile found")
                    else:
                        return False
                    profile_flag = 1
                    return V_relationship_retailer_name,V_relationship_retailer_profile_name,row_index
                else:
                    row_index = row_index + 1
            else:
                self.lo.log_to_file("INFO", "Retailer profile not found")
                return False,False,False

    def check_supplier_profile(self, profie_name, v_doc, row_index):
        so = SeleniumOperations(self.v_task_type, self.v_driver, self.lo)
        self.show_all('//*[@id="form1:table1-nb__xc_c"]', 'all')
        V_relationship_profile_path = '//*[@id="form1:table1"]/table[2]/tbody/tr[' + str(row_index) + ']/td[4]'
        if so.check_exists_by_xpath(V_relationship_profile_path):
            v_supplier_profile_path = '//*[@id="form1:table1:' + str(row_index - 2) + ':commandLink3"]'
            so.click_element_by_xpath(v_supplier_profile_path)
            self.lo.log_to_file("INFO", "In supplier profile")
            so.click_element_by_xpath('//*[@id="form1:table1"]/table[2]/tbody/tr[2]/td[2]/div/a[2]')
            #print('Click on show')
            doc_flag = 0
            map_row_index = 2
            #print(doc_flag)
            while (doc_flag == 0):
                #print('Entry in while')
                v_supplier_profile_doc_path = '//*[@id="form1:table1:0:table2"]/table/tbody/tr[3]/td/table/tbody/tr[' + str(map_row_index) + ']/td[3]'
                if so.check_exists_by_xpath(v_supplier_profile_doc_path):
                    v_supplier_profile_doc = so.get_text_by_xpath(v_supplier_profile_doc_path)
                    if v_supplier_profile_doc == v_doc:
                        v_supplier_doc_status_path = '//*[@id="form1:table1:0:table2"]/table/tbody/tr[3]/td/table/tbody/tr[' + str(map_row_index) + ']/td[9]'
                        if so.check_exists_by_xpath(v_supplier_doc_status_path):
                            v_supplier_doc_status = so.get_text_by_xpath(v_supplier_doc_status_path)
                            if v_supplier_doc_status == 'Active':
                                v_supplier_doc_service_path = '//*[@id="form1:table1:0:table2"]/table/tbody/tr[3]/td/table/tbody/tr[' + str(map_row_index) + ']/td[10]'
                                if so.check_exists_by_xpath(v_supplier_doc_service_path):
                                    v_supplier_doc_service = so.get_text_by_xpath(v_supplier_doc_service_path)
                                    if v_supplier_doc_service == 'FItoService':
                                        self.lo.log_to_file("INFO", "Doc found, status is active, service is correct")
                                        v_supplier_profile_map_version = '//*[@id="form1:table1:0:table2"]/table/tbody/tr[3]/td/table/tbody/tr[' + str(map_row_index) + ']/td[6]'
                                        if so.check_exists_by_xpath(v_supplier_profile_map_version):
                                            v_supplier_profile_map_version_display = so.get_text_by_xpath(v_supplier_profile_map_version)
                                            v_supplier_profile_map_name = '//*[@id="form1:table1:0:table2"]/table/tbody/tr[3]/td/table/tbody/tr[' + str(map_row_index) + ']/td[4]'
                                            if so.check_exists_by_xpath(v_supplier_profile_map_name):
                                                v_supplier_profile_map_name_display = so.get_text_by_xpath(v_supplier_profile_map_name)
                                            else:
                                                return False
                                        else:
                                            return False
                                        return v_supplier_profile_map_version_display, v_supplier_profile_map_name_display,map_row_index
                                        doc_flag = 1
                                    else:
                                        map_row_index = map_row_index + 1
                                        print('Service not FItoService')
                                else:
                                    self.lo.log_to_file("INFO", "No services present in supplier profile")
                                    return False
                            else:
                                map_row_index = map_row_index + 1
                        else:
                            return False
                    else:
                        map_row_index = map_row_index + 1
                else:
                    self.lo.log_to_file("INFO", "No docs added")
                    doc_flag = 1
                    V_new_map_row_index = self.add_capbility(v_doc,profie_name,row_index)
                    return v_supplier_profile_map_version_display, v_supplier_profile_map_name_display, V_new_map_row_index

                profile_flag = 1
        else:
            print('Relationship not found')
            return False


    def search_retailer_map(self,document,retailer_profile_name,supplier_profile_name,supplier_name):
        so = SeleniumOperations(self.v_task_type, self.v_driver, self.lo)
        input_text = '//*[@id="form1:inputText1"]'
        so.send_text_by_xpath(input_text,supplier_name)
        so.click_element_by_xpath('//*[@id="form1:commandButton1"]')
        profile_flag = 0
        row_index = 2
        while(profile_flag == 0):
            V_retailer_relationship_profile_path = '//*[@id="form1:table1"]/table[2]/tbody/tr[' + str(row_index) + ']/td[6]'
            if so.check_exists_by_xpath(V_retailer_relationship_profile_path):
                V_relationship_retailer_profile_name = so.get_text_by_xpath(V_retailer_relationship_profile_path)
                self.lo.log_to_file("INFO", "Relationship found at retailer side")
                if V_relationship_retailer_profile_name == supplier_profile_name:
                    V_relationship_retailer_profile_path = '// *[ @ id = "form1:table1:'+ str(row_index-2) +':commandLink3"]'
                    #so.click_element_by_xpath(V_relationship_retailer_profile_path)
                    if so.check_exists_by_xpath(V_relationship_retailer_profile_path):
                        so.click_element_by_xpath(V_relationship_retailer_profile_path)
                        self.lo.log_to_file("INFO", "In retailer profile")
                        V_retailer_profile_show_path = '//*[@id="form1:table1"]/table[2]/tbody/tr[2]/td[2]/div/a[2]'
                        if so.check_exists_by_xpath(V_retailer_profile_show_path):
                            so.click_element_by_xpath('//*[@id="form1:table1"]/table[2]/tbody/tr[2]/td[2]/div/a[2]')
                            map_flag = 0
                            map_row_index = 2
                            while(map_flag == 0):
                                V_retailer_doc_path = '//*[@id="form1:table1:0:table2"]/table/tbody/tr[3]/td/table/tbody/tr[' + str(map_row_index) + ']/td[3]'
                                if so.check_exists_by_xpath(V_retailer_doc_path):
                                    V_retailer_document = so.get_text_by_xpath(V_retailer_doc_path)
                                    if V_retailer_document == document:
                                        self.lo.log_to_file("INFO", V_retailer_document)
                                        print("Retailer Document = " +V_retailer_document)
                                        self.lo.log_to_file("INFO", "Doc Found")
                                        V_retailer_doc_status_path = '//*[@id="form1:table1:0:table2"]/table/tbody/tr[3]/td/table/tbody/tr[' + str(map_row_index) + ']/td[9]'
                                        if so.check_exists_by_xpath(V_retailer_doc_status_path):
                                            V_reatiler_doc_status = so.get_text_by_xpath(V_retailer_doc_status_path)
                                            if V_reatiler_doc_status == 'Active':
                                                self.lo.log_to_file("INFO", "Status is Active")
                                                print("Retailer Dcoument Status = " + V_reatiler_doc_status)
                                                V_retailer_doc_service_path = '//*[@id="form1:table1:0:table2"]/table/tbody/tr[3]/td/table/tbody/tr[' + str(map_row_index) + ']/td[10]'
                                                if so.check_exists_by_xpath(V_retailer_doc_service_path):
                                                    V_reatiler_doc_service = so.get_text_by_xpath(V_retailer_doc_service_path)
                                                    if V_reatiler_doc_service == 'CommerceEDIfromService' or V_reatiler_doc_service == 'CommerceEDItoService' or V_reatiler_doc_service == 'B2BfromService' or V_reatiler_doc_service == 'B2BtoService':
                                                        self.lo.log_to_file("INFO", "Services are correct")
                                                        print("Retailer doc service = " + V_reatiler_doc_service)
                                                        if(document == '850'):
                                                            V_retailer_doc_version_path = '//*[@id="form1:table1:0:table2"]/table/tbody/tr[3]/td/table/tbody/tr[' + str(map_row_index) + ']/td[8]'
                                                            if so.check_exists_by_xpath(V_retailer_doc_version_path):
                                                                V_retailer_doc_version = so.get_text_by_xpath(V_retailer_doc_version_path)
                                                                self.lo.log_to_file("INFO","Retailer version found")
                                                                print("Retailer map version for 850 =  " + V_retailer_doc_version)
                                                                so.click_element_by_xpath('//*[@id="form1:panelPage1"]/table/tbody/tr[3]/td/table/tbody/tr/td[2]/table/tbody/tr/td[3]/a')
                                                                input_text = '//*[@id="form1:inputText1"]'
                                                                so.send_text_by_xpath(input_text, supplier_name)
                                                                so.click_element_by_xpath('//*[@id="form1:commandButton1"]')
                                                                supp_row_index = row_index - 2
                                                                V_relationship_supplier_name_path = '//*[@id="form1:table1:' + str(supp_row_index) + ':commandLink2"]'
                                                                so.click_element_by_xpath(V_relationship_supplier_name_path)
                                                                so.click_element_by_xpath('//*[@id="form1:panelPage1"]/table/tbody/tr[3]/td/table/tbody/tr/td[2]/table/tbody/tr/td[3]/a')
                                                                return V_retailer_document, V_reatiler_doc_status, V_reatiler_doc_service, V_retailer_doc_version


                                                            else:
                                                                self.lo.log_to_file("INFO","Versions not present")
                                                                return False
                                                        else:
                                                            V_retailer_doc_version_path = '//*[@id="form1:table1:0:table2"]/table/tbody/tr[3]/td/table/tbody/tr[' + str(map_row_index) + ']/td[7]'
                                                            if so.check_exists_by_xpath(V_retailer_doc_version_path):
                                                                V_retailer_doc_version = so.get_text_by_xpath(V_retailer_doc_version_path)
                                                                self.lo.log_to_file("INFO", "Retailer version found")
                                                                print("Retailer map version for" + document+ " = " + V_retailer_doc_version)
                                                                supp_row_index = row_index - 2
                                                                V_relationship_supplier_name_path = '//*[@id="form1:table1:' + str(supp_row_index) + ':commandLink2"]'
                                                                so.click_element_by_xpath(V_relationship_supplier_name_path)
                                                                so.click_element_by_xpath('//*[@id="form1:panelPage1"]/table/tbody/tr[3]/td/table/tbody/tr/td[2]/table/tbody/tr/td[3]/a')
                                                                return V_retailer_document,V_reatiler_doc_status,V_reatiler_doc_service,V_retailer_doc_version
                                                            else:
                                                                self.lo.log_to_file("INFO", "Versions not present")
                                                                return False
                                                        map_flag = 1
                                                    else:
                                                        map_row_index = map_row_index + 1
                                                else:
                                                    self.lo.log_to_file("INFO","No services present in retailer profile")
                                                    return False
                                            else:
                                                map_row_index = map_row_index + 1
                                        else:
                                            return False
                                    else:
                                        map_row_index = map_row_index + 1
                                else:
                                    self.lo.log_to_file("INFO", "No documents present in retailer profile")
                                    map_flag = 1
                                    return False
                        else:
                            return False
                    else:
                        return False
                    profile_flag = 1
                else:
                    row_index = row_index + 1
            else:
                return false

    def add_capbility(self,v_doc,profie_name,row_index):
        so = SeleniumOperations(self.v_task_type, self.v_driver, self.lo)
        v_add_capability_path = '//*[@id="form1:table1:0:table2:addExistingCapability"]'

        print(self.v_driver.current_window_handle)
        if so.check_exists_by_xpath(v_add_capability_path):
            so.click_element_by_xpath(v_add_capability_path)
            time.sleep(1)
            self.v_driver.switch_to.window(self.v_driver.window_handles[-1])
            time.sleep(1)
            self.v_driver.switch_to.frame(0)
            time.sleep(1)
            self.lo.log_to_file("INFO", "Adding capability")
            v_search_existing_path = '//*[@id="form1:showDetailHeader1__xc_"]'
            if so.check_exists_by_xpath(v_search_existing_path):
                so.click_element_by_xpath(v_search_existing_path)
                so.send_text_by_xpath('//*[@id="form1:inputText3"]', 'FItoService')
                so.send_text_by_xpath('//*[@id="form1:inputText5"]', '850')
                self.show_all('//*[@id="form1:showOneChoice2"]', '0')
                self.lo.log_to_file("INFO", "Data sent")
                so.click_element_by_xpath('//*[@id="form1:commandButton2"]')
                self.lo.log_to_file("INFO", "Clicked on search")
                capability_row_index = 0
                capability_flag = 0
                capability_2013 = 'Standard NetSuite Sales Order RSX Order v7.4 Governance (LEGACY SIP)'
                capability_2015 = 'Standard NetSuite Sales Order RSX Order v7.4 Governance (LEGACY SIP) | v3'
                capability_2017_1 = 'Standard NetSuite Sales Order RSX Order v7.7 Governance_v1'
                capability_2017 = 'Standard NetSuite Sales Order RSX Order v7.4 (2017)'
                while (capability_flag == 0):
                    self.lo.log_to_file("INFO", "Inside while")
                    v_choose_capbability_path = '// *[ @ id = "form1:table1:' + str(capability_row_index) + ':outputText6"]'
                    if so.check_exists_by_xpath(v_choose_capbability_path):
                        self.lo.log_to_file("INFO", "Inside If1")
                        v_capability_name = so.get_text_by_xpath(v_choose_capbability_path)
                        print(v_capability_name)
                        if (v_capability_name == capability_2013 or v_capability_name == capability_2015 or v_capability_name == capability_2017 or v_capability_name == capability_2017_1):
                            self.lo.log_to_file("INFO", "Inside If2")
                            so.click_element_by_xpath('//*[@id="form1:table1:' + str(capability_row_index) + ':tableSelectMany1"]')
                            self.lo.log_to_file("INFO", "Tick checkbox")
                            so.click_element_by_xpath('//*[@id="form1:table1:commandButton1"]')
                            self.lo.log_to_file("INFO", "Click choose")
                            capability_flag = 1
                            capability_row_index = capability_row_index + 1
                            self.lo.log_to_file("INFO", "Row++")
                            doc_flag = 0
                            map_row_index = 2
                            while (doc_flag == 0):
                                v_supplier_profile_doc_path = '//*[@id="form1:table1:0:table2"]/table/tbody/tr[3]/td/table/tbody/tr[' + str(map_row_index) + ']/td[3]'
                                if so.check_exists_by_xpath(v_supplier_profile_doc_path):
                                    v_supplier_profile_doc = so.get_text_by_xpath(v_supplier_profile_doc_path)
                                    if v_supplier_profile_doc == v_doc:
                                        v_supplier_doc_status_path = '//*[@id="form1:table1:0:table2"]/table/tbody/tr[3]/td/table/tbody/tr[' + str(map_row_index) + ']/td[9]'
                                        if so.check_exists_by_xpath(v_supplier_doc_status_path):
                                            v_supplier_doc_status = so.get_text_by_xpath(v_supplier_doc_status_path)
                                            if v_supplier_doc_status == 'Active':
                                                v_supplier_doc_service_path = '//*[@id="form1:table1:0:table2"]/table/tbody/tr[3]/td/table/tbody/tr[' + str(map_row_index) + ']/td[10]'
                                                if so.check_exists_by_xpath(v_supplier_doc_service_path):
                                                    v_supplier_doc_service = so.get_text_by_xpath(v_supplier_doc_service_path)
                                                    if v_supplier_doc_service == 'FItoService':
                                                        self.lo.log_to_file("INFO","Doc found, status is active, service is correct")
                                                        v_supplier_profile_map_version = '//*[@id="form1:table1:0:table2"]/table/tbody/tr[3]/td/table/tbody/tr[' + str(map_row_index) + ']/td[6]'
                                                        if so.check_exists_by_xpath(v_supplier_profile_map_version):
                                                            v_supplier_profile_map_version_display = so.get_text_by_xpath(v_supplier_profile_map_version)
                                                            v_supplier_profile_map_name = '//*[@id="form1:table1:0:table2"]/table/tbody/tr[3]/td/table/tbody/tr[' + str(map_row_index) + ']/td[4]'
                                                            if so.check_exists_by_xpath(v_supplier_profile_map_name):
                                                                v_supplier_profile_map_name_display = so.get_text_by_xpath(v_supplier_profile_map_name)
                                                            else:
                                                                return False
                                                        else:
                                                            return False
                                                        return map_row_index
                                                        doc_flag = 1
                                                    else:
                                                        map_row_index = map_row_index + 1
                                                        print('Service not FItoService')
                                                else:
                                                    self.lo.log_to_file("INFO","No services present in supplier profile")
                                                    return False
                                            else:
                                                map_row_index = map_row_index + 1
                                        else:
                                            return False
                                    else:
                                        map_row_index = map_row_index + 1
                        else:
                            self.lo.log_to_file("INFO", "Inside else2")
                            capability_row_index = capability_row_index + 1
                            return False
                    else:
                        return False

    def add_extensions(self,retailer_map_version,supplier_map_version,supplier_map_row_index):
        so = SeleniumOperations(self.v_task_type, self.v_driver, self.lo)
        V_configure_extensions_xpath = '//*[@id="form1:table1:0:table2:' + str(supplier_map_row_index) + ':extensionPopup"]'
        #print(self.v_driver.current_window_handle)
        if so.check_exists_by_xpath(V_configure_extensions_xpath):
            so.click_element_by_xpath(V_configure_extensions_xpath)
            time.sleep(1)
            self.v_driver.switch_to.window(self.v_driver.window_handles[-1])
            time.sleep(1)
            self.v_driver.switch_to.frame(0)
            self.lo.log_to_file("INFO", "Adding extension")
            if(retailer_map_version != supplier_map_version):
                self.lo.log_to_file("INFO", "Adding Map extension")
                v_position = -1
                while(retailer_map_version != supplier_map_version):
                    if retailer_map_version == '7.0':
                        v_position = v_position + 1
                        self.adding_map_extension(LocalElementLocator.PO02,str(v_position))
                        retailer_map_version = '7.2'
                        print('Map Extensions added from 7.0 to 7.2')
                        self.lo.log_to_file("INFO", "Map Extensions added from 7.0 to 7.2")

                    elif retailer_map_version == '7.2':
                        v_position = v_position + 1
                        self.adding_map_extension(LocalElementLocator.PO24,str(v_position))
                        retailer_map_version = '7.4'
                        print('Map Extensions added from 7.2 to 7.4')
                        self.lo.log_to_file("INFO", "Map Extensions added from 7.2 to 7.4")

                    elif retailer_map_version == '7.5':
                        v_position = v_position + 1
                        self.adding_map_extension(LocalElementLocator.PO54,str(v_position))
                        retailer_map_version = '7.4'
                        print('Map Extensions added from 7.5 to 7.4')
                        self.lo.log_to_file("INFO", "Map Extensions added from 7.5 to 7.4")


                    elif retailer_map_version == '7.6':
                        v_position = v_position + 1
                        self.adding_map_extension(LocalElementLocator.PO65,str(v_position))
                        retailer_map_version = '7.5'
                        print('Map Extensions added from 7.6 to 7.5')
                        self.lo.log_to_file("INFO", "Map Extensions added from 7.6 to 7.5")


                    elif retailer_map_version == '7.7':
                        v_position = v_position + 1
                        self.adding_map_extension(LocalElementLocator.PO76,str(v_position))
                        retailer_map_version = '7.6'
                        print('Map Extensions added from 7.7 to 7.6')
                        self.lo.log_to_file("INFO", "Map Extensions added from 7.7 to 7.6")

            if (retailer_map_version == supplier_map_version):
                so.click_element_by_xpath('//*[@id="form1:commandButton2"]')
                self.lo.log_to_file("INFO", "Saving Extension")
                time.sleep(1)
        else:
            self.lo.log_to_file("INFO", "Incorrect path for configure extensions")
            return False

        self.v_driver.switch_to.window(self.v_driver.window_handles[0])



    def adding_map_extension(self, rsx_conversion_map,v_position):
        so = SeleniumOperations(self.v_task_type, self.v_driver, self.lo)
        V_add_extensions_xpath = '//*[@id="form1:table1"]/table[1]/tbody/tr/td/table/tbody/tr/td[1]'
        time.sleep(1)
        if so.check_exists_by_xpath(V_add_extensions_xpath):
            so.click_element_by_xpath(V_add_extensions_xpath)
            V_map_extension_xpath = '//*[@id="form1:table1"]/table/tbody/tr[2]/td/table/tbody/tr[8]/td[1]'
            so.click_element_by_xpath(V_map_extension_xpath)
            self.lo.log_to_file("INFO", "Map Extension added")
            so.click_element_by_xpath('//*[@id="form1:table1:commandButton2"]')
            show_map_extension_xpath = '//*[@id="form1:table1dd' + str(v_position) + '"]'
            if so.check_exists_by_xpath(show_map_extension_xpath):
                so.click_element_by_xpath(show_map_extension_xpath)
                V_map_extension_input_xpath = '//*[@id="form1:table1:' + str(v_position) +':table2:0:outputText22"]'
                if so.check_exists_by_xpath(V_map_extension_input_xpath):
                    so.send_text_by_xpath(V_map_extension_input_xpath,rsx_conversion_map)
                    self.lo.log_to_file("INFO", "Conversion Map added")
                    return True
                else:
                    self.lo.log_to_file("INFO", "Incorrect path for inout text")
                    return False
            else:
                self.lo.log_to_file("INFO","Incorrect path for show map extension")
                return False
        else:
            self.lo.log_to_file("INFO", "Incorrect path for add extensions xpath")
            return False



    def execute_main(self):
        v_start_time = time.time()
        self.lo.log_to_file("INFO", "Login in to DC4 Pre-Prod")
        lg = Login(self.v_task_type, self.v_driver, self.v_input_wb, self.lo)
        so = SeleniumOperations(self.v_task_type,self.v_driver,self.lo)
        rf = ReportFileUtility(self.v_task_type)
        SeleniumOperations(self.v_task_type, self.v_driver, self.lo)
        output_sheet = self.v_input_wb.get_sheet_by_name('Output')
        input_sheet = self.v_input_wb.get_sheet_by_name('Input')
        input_sheet_ex = ExcelOperations(self.v_task_type, input_sheet)
        output_sheet_ex = ExcelOperations(self.v_task_type, output_sheet)
        self.v_driver.maximize_window()
        Max_row = input_sheet.max_row
        current_row = 2
        #Max_row = output_sheet.max_row
        #current_row_output = 2


        lg.login("DC4 PreProd")

        while(current_row <= Max_row):
            v_TPID = input_sheet.cell(row = current_row, column = 1).value
            v_Doc_850 = input_sheet.cell(row = current_row, column = 2).value
            if (v_Doc_850 == 'Y'):
                v_doc = '850'
                V_supplier_doc_version = '7.4'
            dc = DC4_Utility(self.v_task_type, self.v_driver, self.lo)
            dc.company_search_by_TPID(v_TPID)
            self.lo.log_to_file("INFO", "Searching Supplier Data")
            V_Supplier_Profile_name = so.get_text_by_xpath('//*[@id="table2"]/table/tbody/tr/td/table/tbody/tr[2]/td[2]')
            V_Supplier_name = so.get_text_by_xpath('//*[@id="table2:0:commandLink4"]')
            print("Supplier Name = " + V_Supplier_name)
            output_sheet_ex.set_value(current_row, 1,V_Supplier_name)
            print("Supplier Profile = " + V_Supplier_Profile_name)
            so.click_element_by_xpath('//*[@id="table2:0:commandLink4"]')
            so.click_element_by_xpath('//*[@id="form1:panelPage1"]/table/tbody/tr[3]/td/table/tbody/tr/td[2]/table/tbody/tr/td[3]/a')
            self.lo.log_to_file("INFO", "Searching Retailer Data")
            V_retailer_name, V_retailer_profile, V_supplier_row_index = self.search_retailer_profile(V_Supplier_Profile_name)
            #print(V_retailer_name, V_retailer_profile, V_supplier_row_index)
            if V_retailer_name == V_retailer_profile == V_supplier_row_index == False:
                print('No relationship found')
                self.lo.log_to_file("INFO", "No relationship found, going to next row")
                output_sheet_ex.set_value(current_row, 4, 'Relationship not found')
                current_row = current_row + 1
            else:
                print("Retailer Name = " + V_retailer_name)
                output_sheet_ex.set_value(current_row, 2, V_retailer_name)
                print("Retailer Profile = " + V_retailer_profile)
                self.lo.log_to_file("INFO", "Searching Retailer Map Versions")
                V_ret_document, V_ret_doc_status, V_ret_doc_service, V_ret_doc_version = self.search_retailer_map(v_doc,V_retailer_profile,V_Supplier_Profile_name,V_Supplier_name)
                V_supplier_map_version, V_supplier_profile_map_name, V_supplier_map_row_index = self.check_supplier_profile(V_Supplier_Profile_name, v_doc,V_supplier_row_index)
                print("Supplier Map version = " + V_supplier_map_version)
                print("Supplier Map Name" + V_supplier_profile_map_name)
                output_sheet_ex.set_value(current_row, 3, V_ret_document)
                self.v_input_wb.save(AppConstants.INPUT_FILE_PATH)
                self.lo.log_to_file("INFO", "Maps and version found")
                V_send_suuplier_map_index = int(V_supplier_map_row_index)
                V_send_suuplier_map_index = V_send_suuplier_map_index - 2
                status = self.add_extensions(V_ret_doc_version, V_supplier_doc_version, V_send_suuplier_map_index)
                #print(status)
                if status == 'None':
                    output_sheet_ex.set_value(current_row, 4, 'Success')
                else:
                    output_sheet_ex.set_value(current_row, 4, 'Fail')
                current_row = current_row + 1

        self.v_driver.close()
        v_end_time = time.time()
        rf.update_sheet(self.v_username, 2, math.floor(v_end_time - v_start_time), str(datetime.date.today()))





